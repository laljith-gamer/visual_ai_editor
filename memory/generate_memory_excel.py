"""Generate a day-wise Excel workbook from the project memory files.

Reads every memory/*.md file, extracts its date (from filename or content),
its title/summary, and writes a formatted .xlsx organized day-by-day.
"""

import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MEMORY_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "memory")
OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "memory_files_daywise.xlsx")

# --- Style constants -------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATE_FILL = PatternFill("solid", fgColor="2E75B6")
DATE_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
CORE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATE_RE = re.compile(r"(20\d{2}-\d{2}-\d{2})")


def extract_date(filename: str, content: str) -> str:
    """Find the date for a memory file: prefer filename, then first line date."""
    m = DATE_RE.search(filename)
    if m:
        return m.group(1)
    # fall back to first date found in the body (e.g. core files)
    m = DATE_RE.search(content)
    if m:
        return m.group(1)
    return "undated"


def first_heading(content: str) -> str:
    for line in content.splitlines():
        s = line.strip()
        if s.startswith("# "):
            return s.lstrip("# ").strip()
    return "(no title)"


def first_paragraph(content: str) -> str:
    """Grab the first non-heading, non-empty, non-quote line as a summary."""
    for line in content.splitlines():
        s = line.strip()
        if not s:
            continue
        if s.startswith("#"):
            continue
        if s.startswith(">"):
            continue
        if s.startswith("```"):
            continue
        return s
    return ""


def files_changed_guess(content: str, filename: str) -> str:
    """Collect code-ish file paths mentioned in the doc."""
    paths = re.findall(r"`([a-zA-Z0-9_\-./]+(?:\.ts|\.tsx|\.js|\.mjs|\.css|\.json|\.md))`", content)
    seen = []
    for p in paths:
        if p not in seen and not p.startswith("memory/"):
            seen.append(p)
    return ", ".join(seen[:8]) + (" ..." if len(seen) > 8 else "")


def is_core_file(filename: str) -> bool:
    return filename in {
        "INDEX.md", "PROJECT_STATE.md", "CHANGELOG.md", "DECISIONS.md",
        "CONSTRAINTS.md", "ROADMAP.md", "TODO.md", "MEMORY_FILES_ORGANIZED_BY_DATE.md",
    }


def main():
    rows = []
    for fn in sorted(os.listdir(MEMORY_DIR)):
        if not fn.endswith(".md"):
            continue
        path = os.path.join(MEMORY_DIR, fn)
        try:
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
        except Exception:
            continue
        date = extract_date(fn, content)
        title = first_heading(content)
        summary = first_paragraph(content)
        files = files_changed_guess(content, fn)
        core = is_core_file(fn)
        rows.append({
            "date": date,
            "file": fn,
            "title": title,
            "summary": summary,
            "files_changed": files,
            "core": core,
        })

    # sort: dated files chronological (undated last), then core files last
    def sort_key(r):
        if r["date"] == "undated":
            return ("9", r["file"])
        return (r["date"], r["file"])
    rows.sort(key=sort_key)

    wb = Workbook()

    # ---- Sheet 1: Day-wise ------------------------------------------------
    ws = wb.active
    ws.title = "Day-wise"
    headers = ["Date", "Day", "File", "Title", "Summary", "Files Changed", "Type"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    current_date = None
    for r in rows:
        d = r["date"]
        if d != current_date:
            current_date = d
            # date divider row
            ws.append([d, "", "", "", "", "", ""])
            ridx = ws.max_row
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=ridx, column=c)
                cell.fill = DATE_FILL
                cell.font = DATE_FONT
                cell.border = BORDER
        try:
            dayname = datetime.strptime(d, "%Y-%m-%d").strftime("%A")
        except Exception:
            dayname = ""
        ws.append([
            d,
            dayname,
            r["file"],
            r["title"],
            r["summary"],
            r["files_changed"],
            "Core" if r["core"] else "Dated",
        ])
        ridx = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=ridx, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
            if r["core"]:
                cell.fill = CORE_FILL

    widths = [12, 12, 42, 34, 70, 46, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---- Sheet 2: Per-day counts -----------------------------------------
    ws2 = wb.create_sheet("Summary by Day")
    ws2.append(["Date", "Day", "Files", "Core Files", "Dated Files"])
    for c in range(1, 6):
        cell = ws2.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    from collections import defaultdict
    by_day = defaultdict(lambda: {"total": 0, "core": 0, "dated": 0})
    for r in rows:
        d = r["date"]
        by_day[d]["total"] += 1
        if r["core"]:
            by_day[d]["core"] += 1
        else:
            by_day[d]["dated"] += 1
    for d in sorted(by_day.keys()):
        try:
            dayname = datetime.strptime(d, "%Y-%m-%d").strftime("%A")
        except Exception:
            dayname = ""
        info = by_day[d]
        ws2.append([d, dayname, info["total"], info["core"], info["dated"]])
        ridx = ws2.max_row
        for c in range(1, 6):
            ws2.cell(row=ridx, column=c).border = BORDER
            ws2.cell(row=ridx, column=c).alignment = CENTER
    for i, w in enumerate([14, 12, 10, 12, 12], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print(f"Rows: {len(rows)} files across {len(by_day)} days")


if __name__ == "__main__":
    main()
